from PyQt5 import QtWidgets
from PyQt5.QtWidgets import *
from chatterbot import ChatBot
from chatterbot.trainers import ChatterBotCorpusTrainer
from PyQt5.QtGui import QMovie
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import TextToSpeach as tts
import SoundRecognizer as soundRecognizer
import  bot
from ui import ui
import threading
import time
import re
from CommandModules import Web

from openpyxl import load_workbook

class Main(QMainWindow, ui.Ui_MainWindow):

    def __init__(self):
         super().__init__()
         self.setupUi(self)
         #線程們
         self.SpeackAniThread = SpeackAniWorkThread()
         self.SpeackSoundThread = SpeackSoundWorkThread()
         self.CommandModuleThread = CommandModuleThread()
         self.ListenThread = ListenSoundWorkThread()
         self.ChatBotThread = ChatBotWorkThread()
         self.ChatBotThread.init()
         self.ListenThread.setTerminationEnabled(True)
         self.CommandModuleThread.setTerminationEnabled(True)
         self.SpeackSoundThread.setTerminationEnabled(True)
         self.SpeackAniThread.setTerminationEnabled(True)
         self.ChatBotThread.setTerminationEnabled(True)
         #terminate enable
         self.enterbtn.clicked.connect(self.onEnterClicked)
         self.speackbtn.clicked.connect(self.onSpeackClicked)


         self.ListenThread.trigger.connect(self.onUserSpeack)
         self.CommandModuleThread.trigger.connect(self.onResultTextRecive)
         self.SpeackAniThread.trigger.connect(self.playTalkAnimation)
         self.ChatBotThread.trigger.connect(self.showText)

    def onSpeackClicked(self):
        print("開始聆聽...")
        # kill thread-----
        self.ListenThread.stop()
        self.CommandModuleThread.stop()
        self.SpeackSoundThread.stop()
        self.SpeackAniThread.stop()
        # kill thread-----
        self.ListenThread.start()


    def onEnterClicked(self):
        # kill thread-----
        self.ListenThread.stop()
        self.CommandModuleThread.stop()
        self.SpeackSoundThread.stop()
        self.SpeackAniThread.stop()
        #kill thread-----
        input = self.lineEdit.text()
        self.CommandModuleThread.setWord(input)
        self.CommandModuleThread.start()


    #收到結果 要在判斷是 指令的回應 還是一班聊天要再透過chatbot取得答案
    def onResultTextRecive(self,str):
        print("::onResultTextRecive收到::")
        print(str)
        word = str.split('/')
        commandType = word[0]
        #有指令
        # command + function + callback
        if commandType=="command":
            command = word[1]
            callback = word[2]
            print("要執行:"+command)

            print("call back要說:"+callback)
            print(exec("result=" + str(command)))
        else:
            self.ChatBotThread.setWord(word[1])
            self.ChatBotThread.start()


    #收到使用者的文字
    def onUserSpeack(self,word):
        self.lineEdit.setText(word)
        self.CommandModuleThread.setWord(word)
        self.CommandModuleThread.start()



    def playTalkAnimation(self,str):
        #收到0是開始 1是結束
        #print("收到:"+str)
        if str=="0":
            self.playGif()
        elif str=="1":
            self.hideGif()

    #顯示結果 會順便講話
    def showText(self,text):
        print("::showtext::"+text)
        #播放動畫
        self.SpeackAniThread.setTime(3)
        self.SpeackAniThread.start()
        #thread 執行 tts
        self.resulttext.setPlainText(text)
        self.SpeackSoundThread.setWord(text)
        self.SpeackSoundThread.start()


#說話動畫線程
class SpeackAniWorkThread(QThread):
    # 自定义信号对象。参数str就代表这个信号可以传一个字符串
    trigger = pyqtSignal(str)
    def __int__(self):
        # 初始化函数
        super(SpeackAniWorkThread, self).__init__()

    def run(self):
        #重写线程执行的run函数
        #触发自定义信号
        self.trigger.emit(str(0))
        for i in range(self.waittime):
            time.sleep(1)
            # 通过自定义信号把待显示的字符串传递给槽函数
            if i>=self.waittime-1:
                self.trigger.emit(str(1))

    def setTime(self,time):
        self.waittime = time
    def stop(self):
        self.terminate()

#說話線程
class SpeackSoundWorkThread(QThread):
    def __int__(self):
        # 初始化函数
        super(SpeackAniWorkThread, self).__init__()

    def run(self):
        tts.Speack(self.word)

    def setWord(self,word):
        print("set word:"+word)
        self.word = word

    def stop(self):
        self.terminate()

class ListenSoundWorkThread(QThread):
    trigger = pyqtSignal(str)
    def __int__(self):
        # 初始化函数
        super(SpeackAniWorkThread, self).__init__()

    def run(self):
        data = soundRecognizer.Listen()
        print("結果:")
        print(data)
        self.trigger.emit(str(data))
    def stop(self):
        self.terminate()

#指令線程
class CommandModuleThread(QThread):
    trigger = pyqtSignal(str)
    def __init__(self):
        super(CommandModuleThread,self).__init__()
        wb = load_workbook('Commands/commandlist.xlsx')
        ws = wb.worksheets[0]
        self.commandList = dict()
        for i in range(2, ws.max_column + 1):
            self.commandList[ws.cell(row=1, column=i).value] = [ws.cell(row=3, column=i).value,
                                                                ws.cell(row=4, column=i).value]
        print(self.commandList)
        print("Load command...done")

    def setWord(self,word):
        #print("set command:"+word)
        self.command = word

    def run(self):
        talk = self.command.lower()
        print(talk)
        hadcommand = False
        _commandindex = 0
        for t in self.commandList.keys():
            if re.findall(t, talk):
                # 執行excel下的指令
                # exec(self.commandList[t][0])
                # 講出callback
                print(self.commandList[t][0])
                _commandindex = t
                hadcommand = True
        if hadcommand is True:
            #有指令
            print(":conmandModule emmit had command:")
            print("指令是:")
            print(str(self.commandList[_commandindex][0]))
            print("Call back世說:")
            print(self.commandList[_commandindex][1])
            #command + function + callback

            self.trigger.emit("command/"+str(self.commandList[_commandindex][0])+"/"+self.commandList[_commandindex][1])
        else:
            print(":conmandModule emmit:")
            self.trigger.emit("normal/"+str(talk))

    def stop(self):
        self.terminate()

#ChatBot線程
class ChatBotWorkThread(QThread):
    trigger = pyqtSignal(str)
    def __int__(self):
        # 初始化函数
        super(ChatBotWorkThread, self).__init__()

    def init(self):
        print("init...?")
        self.ChatBot = bot.ChatBotClass()
        self.ChatBot.init()
        self.ChatBot.hellowGreed()
        result = self.ChatBot.getResponse("你好")
        print("::chatbotModule init::" + result)

    def setWord(self,word):
        self.word = word

    def run(self):
        print("::chatbotModule:: begin")
        result = self.ChatBot.getResponse(self.word)
        print("::chatbotModule::"+result)
        self.trigger.emit(str(result))

    def stop(self):
        self.terminate()

if __name__ == '__main__':
    import sys
    app = QtWidgets.QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())